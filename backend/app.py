"""
NCEMC Site Visit Check-In System — API Backend
================================================
Deployed on Render. Frontend deployed separately on Vercel.

Auth uses signed tokens (itsdangerous) sent via Authorization header.
CORS enabled for the frontend origin.
"""

import os
import json
import sqlite3
import hashlib
import secrets
import functools
import threading
from datetime import datetime, date, timedelta
from collections import defaultdict
from io import BytesIO
from queue import Queue

from flask import Flask, request, jsonify, Response, stream_with_context, g
from flask_cors import CORS
from itsdangerous import URLSafeTimedSerializer

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "visitors.db"))
EXCEL_PATH = os.environ.get("EXCEL_PATH", os.path.join(BASE_DIR, "Information Note Master.xlsm"))

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", secrets.token_hex(32))

FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")

CORS(app,
     origins=[FRONTEND_URL, "http://localhost:3000", "http://127.0.0.1:3000"],
     allow_headers=["Content-Type", "Authorization"],
     expose_headers=["Content-Disposition"],
     supports_credentials=False)

# SSE clients
sse_clients = []
sse_lock = threading.Lock()

# Rate limiting
rate_limit_store = defaultdict(list)
RATE_LIMIT = 30
RATE_WINDOW = 60


def check_rate_limit():
    ip = request.remote_addr or "unknown"
    now = __import__("time").time()
    rate_limit_store[ip] = [t for t in rate_limit_store[ip] if now - t < RATE_WINDOW]
    if len(rate_limit_store[ip]) >= RATE_LIMIT:
        return True
    rate_limit_store[ip].append(now)
    return False


# ---------------------------------------------------------------------------
# Auth — Token-based
# ---------------------------------------------------------------------------
token_serializer = None

def get_serializer():
    global token_serializer
    if token_serializer is None:
        token_serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])
    return token_serializer


def generate_auth_token():
    return get_serializer().dumps({"auth": True}, salt="staff-auth")


def verify_auth_token(token, max_age=86400):
    try:
        data = get_serializer().loads(token, salt="staff-auth", max_age=max_age)
        return data.get("auth") is True
    except Exception:
        return False


def token_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        tok = None
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            tok = auth_header[7:]
        if not tok:
            tok = request.args.get("token")
        if not tok or not verify_auth_token(tok):
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Password helpers
# ---------------------------------------------------------------------------
def hash_password(pw):
    salt = os.urandom(16).hex()
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 100000).hex()
    return f"{salt}:{h}"


def verify_password(pw, stored):
    if ":" not in stored:
        return False
    salt, h = stored.split(":", 1)
    return hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 100000).hex() == h


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db:
        db.close()


def audit(action, detail, actor="system"):
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT INTO audit_log (action, detail, actor) VALUES (?, ?, ?)",
        (action, detail, actor)
    )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# DB init & Excel import
# ---------------------------------------------------------------------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS config (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS visitors (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            name     TEXT NOT NULL,
            company  TEXT NOT NULL DEFAULT '',
            phone    TEXT DEFAULT '',
            approved INTEGER DEFAULT 1,
            added_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(name, company)
        );
        CREATE TABLE IF NOT EXISTS companies (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS sites (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS company_sites (
            company TEXT NOT NULL,
            site_id INTEGER NOT NULL,
            UNIQUE(company, site_id),
            FOREIGN KEY (site_id) REFERENCES sites(id)
        );
        CREATE TABLE IF NOT EXISTS reasons (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS checkins (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            visitor_id      INTEGER NOT NULL,
            site            TEXT,
            estimated_time  TEXT,
            reason          TEXT,
            notes           TEXT DEFAULT '',
            time_in         TEXT NOT NULL,
            time_out        TEXT,
            uar_number      TEXT DEFAULT '',
            ticket_closed   TEXT DEFAULT '',
            alarms_clear    TEXT DEFAULT '',
            alarm_details   TEXT DEFAULT '',
            complete        INTEGER DEFAULT 0,
            is_walkin       INTEGER DEFAULT 0,
            date_stamp      TEXT NOT NULL,
            FOREIGN KEY (visitor_id) REFERENCES visitors(id)
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp  TEXT DEFAULT (datetime('now','localtime')),
            action     TEXT NOT NULL,
            detail     TEXT,
            actor      TEXT DEFAULT 'system'
        );
    """)

    cur = conn.cursor()
    cur.execute("SELECT value FROM config WHERE key = 'staff_password'")
    if not cur.fetchone():
        default_pw = secrets.token_urlsafe(8)
        cur.execute(
            "INSERT INTO config (key, value) VALUES ('staff_password', ?)",
            (hash_password(default_pw),)
        )
        conn.commit()
        print(f"\n  ** Default staff password: {default_pw} **")
        print(f"  ** Change it via Admin > Settings after first login **\n")
        cur.execute(
            "INSERT INTO audit_log (action, detail) VALUES (?, ?)",
            ("system_init", "Default staff password generated")
        )
        conn.commit()

    conn.execute("INSERT OR IGNORE INTO reasons (name) VALUES ('Other')")
    conn.commit()
    conn.close()


def import_from_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"[INFO] Excel file not found at {EXCEL_PATH} — skipping import.")
        return

    conn = sqlite3.connect(DB_PATH)
    if conn.execute("SELECT COUNT(*) FROM visitors").fetchone()[0] > 0:
        print("[INFO] Database already has visitor data — skipping Excel import.")
        conn.close()
        return

    try:
        import openpyxl
    except ImportError:
        print("[WARN] openpyxl not installed — cannot import Excel data.")
        conn.close()
        return

    print("[INFO] Importing data from Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Lists"]
    cur = conn.cursor()

    visitor_count = 0
    companies_seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
        name, company, phone = row
        if name:
            name = str(name).strip()
            company = str(company).strip() if company else ""
            phone = str(phone).strip() if phone else ""
            cur.execute(
                "INSERT OR IGNORE INTO visitors (name, company, phone, approved) VALUES (?, ?, ?, 1)",
                (name, company, phone)
            )
            visitor_count += 1
            if company:
                companies_seen.add(company)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=17, max_col=17, values_only=True):
        if row[0]:
            companies_seen.add(str(row[0]).strip())

    for c in sorted(companies_seen):
        cur.execute("INSERT OR IGNORE INTO companies (name) VALUES (?)", (c,))

    site_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10, values_only=True):
        if row[0]:
            cur.execute("INSERT OR IGNORE INTO sites (name) VALUES (?)", (str(row[0]).strip(),))
            site_count += 1

    reason_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8, values_only=True):
        if row[0]:
            cur.execute("INSERT OR IGNORE INTO reasons (name) VALUES (?)", (str(row[0]).strip(),))
            reason_count += 1

    mapping_count = 0
    for col_idx in range(19, 61):
        company_name = ws.cell(row=1, column=col_idx).value
        if not company_name:
            continue
        company_name = str(company_name).strip()
        for row_idx in range(2, ws.max_row + 1):
            site_name = ws.cell(row=row_idx, column=col_idx).value
            if site_name:
                site_name = str(site_name).strip()
                cur.execute("SELECT id FROM sites WHERE name = ?", (site_name,))
                site_row = cur.fetchone()
                if site_row:
                    cur.execute(
                        "INSERT OR IGNORE INTO company_sites (company, site_id) VALUES (?, ?)",
                        (company_name, site_row[0])
                    )
                    mapping_count += 1

    cur.execute(
        "INSERT INTO audit_log (action, detail) VALUES (?, ?)",
        ("excel_import", f"{visitor_count} visitors, {site_count} sites, {reason_count} reasons, {mapping_count} mappings")
    )
    conn.commit()
    conn.close()
    wb.close()
    print(f"[INFO] Imported {visitor_count} visitors, {len(companies_seen)} companies, "
          f"{site_count} sites, {reason_count} reasons, {mapping_count} company-site mappings.")


# ---------------------------------------------------------------------------
# SSE
# ---------------------------------------------------------------------------
def notify_dashboard(event_type, data):
    message = f"event: {event_type}\ndata: {json.dumps(data)}\n\n"
    dead = []
    with sse_lock:
        for q in sse_clients:
            try:
                q.put_nowait(message)
            except Exception:
                dead.append(q)
        for q in dead:
            sse_clients.remove(q)


# ---------------------------------------------------------------------------
# Routes — Auth
# ---------------------------------------------------------------------------
@app.route("/api/login", methods=["POST"])
def api_login():
    if check_rate_limit():
        return jsonify({"error": "Too many attempts. Wait a minute."}), 429

    data = request.get_json()
    password = data.get("password", "")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT value FROM config WHERE key = 'staff_password'")
    row = cur.fetchone()
    conn.close()

    if row and verify_password(password, row[0]):
        token = generate_auth_token()
        audit("staff_login", f"IP: {request.remote_addr}", "staff")
        return jsonify({"success": True, "token": token})

    return jsonify({"error": "Invalid password"}), 403


@app.route("/api/logout", methods=["POST"])
@token_required
def api_logout():
    audit("staff_logout", f"IP: {request.remote_addr}", "staff")
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Routes — Public (visitor-facing, rate-limited)
# ---------------------------------------------------------------------------
@app.route("/api/companies")
def api_companies():
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429
    db = get_db()
    rows = db.execute("SELECT name FROM companies ORDER BY name").fetchall()
    return jsonify([r["name"] for r in rows])


@app.route("/api/visitors-by-company")
def api_visitors_by_company():
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429

    company = request.args.get("company", "").strip()
    query = request.args.get("q", "").strip().lower()
    if not company:
        return jsonify([])

    db = get_db()
    rows = db.execute(
        "SELECT id, name, company, phone, approved FROM visitors WHERE company = ? ORDER BY name",
        (company,)
    ).fetchall()

    results = []
    for r in rows:
        if not query or query in r["name"].lower():
            results.append({
                "id": r["id"], "name": r["name"],
                "company": r["company"], "phone": r["phone"],
                "approved": r["approved"],
            })

    results.sort(key=lambda x: (0 if x["name"].lower().startswith(query) else 1, x["name"]))
    return jsonify(results[:25])


@app.route("/api/sites/<path:company>")
def api_sites_for_company(company):
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429
    db = get_db()
    rows = db.execute("""
        SELECT s.name FROM sites s
        JOIN company_sites cs ON cs.site_id = s.id
        WHERE cs.company = ?
        ORDER BY s.name
    """, (company,)).fetchall()
    if not rows:
        rows = db.execute("SELECT name FROM sites ORDER BY name").fetchall()
    return jsonify([r["name"] for r in rows])


@app.route("/api/reasons")
def api_reasons():
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429
    db = get_db()
    rows = db.execute("SELECT name FROM reasons ORDER BY name").fetchall()
    return jsonify([r["name"] for r in rows])


@app.route("/api/checkin", methods=["POST"])
def api_checkin():
    if check_rate_limit():
        return jsonify({"error": "Too many requests. Please wait."}), 429

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    site = (data.get("site") or "").strip()
    estimated_time = (data.get("estimated_time") or "").strip()
    reason = (data.get("reason") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not site:
        return jsonify({"error": "Site is required"}), 400
    if not reason:
        return jsonify({"error": "Reason is required"}), 400

    db = get_db()
    is_walkin = False
    visitor_id = data.get("visitor_id")

    if visitor_id:
        visitor = db.execute(
            "SELECT id, name, company, phone, approved FROM visitors WHERE id = ?",
            (visitor_id,)
        ).fetchone()
        if not visitor:
            return jsonify({"error": "Visitor not found"}), 404
        v_name, v_company, v_phone = visitor["name"], visitor["company"], visitor["phone"]
    else:
        walkin_name = (data.get("walkin_name") or "").strip()
        walkin_company = (data.get("walkin_company") or "").strip()
        walkin_phone = (data.get("walkin_phone") or "").strip()
        if not walkin_name or not walkin_company:
            return jsonify({"error": "Name and company are required"}), 400

        is_walkin = True
        existing = db.execute(
            "SELECT id FROM visitors WHERE name = ? AND company = ?",
            (walkin_name, walkin_company)
        ).fetchone()

        if existing:
            visitor_id = existing["id"]
        else:
            cur = db.execute(
                "INSERT INTO visitors (name, company, phone, approved) VALUES (?, ?, ?, 0)",
                (walkin_name, walkin_company, walkin_phone)
            )
            db.commit()
            visitor_id = cur.lastrowid
            db.execute("INSERT OR IGNORE INTO companies (name) VALUES (?)", (walkin_company,))
            db.commit()

        v_name, v_company, v_phone = walkin_name, walkin_company, walkin_phone

    today = date.today().isoformat()
    already = db.execute(
        "SELECT id FROM checkins WHERE visitor_id = ? AND time_out IS NULL AND date_stamp = ?",
        (visitor_id, today)
    ).fetchone()
    if already:
        return jsonify({"error": "Already checked in today", "checkin_id": already["id"]}), 409

    time_in = datetime.now().strftime("%H:%M")
    cur = db.execute(
        """INSERT INTO checkins
           (visitor_id, site, estimated_time, reason, notes, time_in, date_stamp, is_walkin)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (visitor_id, site, estimated_time, reason, notes, time_in, today, int(is_walkin))
    )
    db.commit()
    checkin_id = cur.lastrowid

    audit("checkin",
          f"{'WALK-IN ' if is_walkin else ''}{v_name} ({v_company}) at {site} — {reason}",
          "visitor")

    notify_dashboard("new_checkin", {
        "id": checkin_id, "visitor_id": visitor_id,
        "visitor_name": v_name, "company": v_company, "phone": v_phone,
        "site": site, "estimated_time": estimated_time, "reason": reason,
        "notes": notes, "time_in": time_in, "is_walkin": is_walkin,
    })

    return jsonify({"success": True, "checkin_id": checkin_id, "time_in": time_in, "visitor_id": visitor_id})


@app.route("/api/checkout", methods=["POST"])
def api_checkout_self():
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429
    data = request.get_json()
    visitor_id = data.get("visitor_id")
    if not visitor_id:
        return jsonify({"error": "No visitor specified"}), 400

    db = get_db()
    today = date.today().isoformat()
    checkin = db.execute(
        "SELECT id FROM checkins WHERE visitor_id = ? AND time_out IS NULL AND date_stamp = ?",
        (visitor_id, today)
    ).fetchone()
    if not checkin:
        return jsonify({"error": "No active check-in found"}), 404

    now = datetime.now().strftime("%H:%M")
    db.execute("UPDATE checkins SET time_out = ? WHERE id = ?", (now, checkin["id"]))
    db.commit()

    visitor = db.execute("SELECT name, company FROM visitors WHERE id = ?", (visitor_id,)).fetchone()
    audit("checkout", f"{visitor['name']} ({visitor['company']}) at {now}", "visitor")
    notify_dashboard("checkout", {
        "id": checkin["id"], "visitor_name": visitor["name"],
        "company": visitor["company"], "time_out": now,
    })
    return jsonify({"success": True, "time_out": now})


@app.route("/api/check-status/<int:visitor_id>")
def api_check_status(visitor_id):
    if check_rate_limit():
        return jsonify({"error": "Rate limited"}), 429
    db = get_db()
    today = date.today().isoformat()
    checkin = db.execute(
        "SELECT id, time_in, site FROM checkins WHERE visitor_id = ? AND time_out IS NULL AND date_stamp = ?",
        (visitor_id, today)
    ).fetchone()
    if checkin:
        return jsonify({"checked_in": True, "checkin_id": checkin["id"],
                        "time_in": checkin["time_in"], "site": checkin["site"]})
    return jsonify({"checked_in": False})


# ---------------------------------------------------------------------------
# Routes — Dashboard (token required)
# ---------------------------------------------------------------------------
@app.route("/api/dashboard/checkins")
@token_required
def api_dashboard_checkins():
    db = get_db()
    date_from = request.args.get("from", date.today().isoformat())
    date_to = request.args.get("to", date_from)

    rows = db.execute("""
        SELECT c.*, v.name as visitor_name, v.company, v.phone, v.approved
        FROM checkins c JOIN visitors v ON v.id = c.visitor_id
        WHERE c.date_stamp >= ? AND c.date_stamp <= ?
        ORDER BY c.date_stamp DESC, c.time_in DESC
    """, (date_from, date_to)).fetchall()

    return jsonify([{
        "id": r["id"], "visitor_id": r["visitor_id"],
        "visitor_name": r["visitor_name"], "company": r["company"],
        "phone": r["phone"], "site": r["site"],
        "estimated_time": r["estimated_time"], "reason": r["reason"],
        "notes": r["notes"], "time_in": r["time_in"], "time_out": r["time_out"],
        "uar_number": r["uar_number"], "ticket_closed": r["ticket_closed"],
        "alarms_clear": r["alarms_clear"], "alarm_details": r["alarm_details"],
        "complete": r["complete"], "is_walkin": r["is_walkin"],
        "approved": r["approved"], "date_stamp": r["date_stamp"],
    } for r in rows])


@app.route("/api/dashboard/update/<int:checkin_id>", methods=["POST"])
@token_required
def api_dashboard_update(checkin_id):
    data = request.get_json()
    db = get_db()
    allowed = ["time_out", "uar_number", "ticket_closed",
               "alarms_clear", "alarm_details", "notes", "complete"]
    updates, values = [], []
    for field in allowed:
        if field in data:
            updates.append(f"{field} = ?")
            values.append(data[field])
    if not updates:
        return jsonify({"error": "Nothing to update"}), 400

    values.append(checkin_id)
    db.execute(f"UPDATE checkins SET {', '.join(updates)} WHERE id = ?", values)
    db.commit()

    row = db.execute("""
        SELECT c.*, v.name as visitor_name, v.company, v.phone, v.approved
        FROM checkins c JOIN visitors v ON v.id = c.visitor_id WHERE c.id = ?
    """, (checkin_id,)).fetchone()

    if row:
        notify_dashboard("update_checkin", {
            "id": row["id"], "visitor_id": row["visitor_id"],
            "visitor_name": row["visitor_name"], "company": row["company"],
            "phone": row["phone"], "site": row["site"],
            "estimated_time": row["estimated_time"], "reason": row["reason"],
            "notes": row["notes"], "time_in": row["time_in"], "time_out": row["time_out"],
            "uar_number": row["uar_number"], "ticket_closed": row["ticket_closed"],
            "alarms_clear": row["alarms_clear"], "alarm_details": row["alarm_details"],
            "complete": row["complete"], "is_walkin": row["is_walkin"],
            "approved": row["approved"],
        })
        audit("update_checkin", f"#{checkin_id} — {row['visitor_name']}", "staff")
    return jsonify({"success": True})


@app.route("/api/dashboard/approve/<int:visitor_id>", methods=["POST"])
@token_required
def api_approve_visitor(visitor_id):
    db = get_db()
    visitor = db.execute("SELECT name, company, phone FROM visitors WHERE id = ?", (visitor_id,)).fetchone()
    if not visitor:
        return jsonify({"error": "Visitor not found"}), 404
    db.execute("UPDATE visitors SET approved = 1 WHERE id = ?", (visitor_id,))
    db.commit()
    audit("approve_visitor", f"{visitor['name']} ({visitor['company']}) added to directory", "staff")
    notify_dashboard("visitor_approved", {"visitor_id": visitor_id, "name": visitor["name"]})
    return jsonify({"success": True, "message": f"{visitor['name']} added to directory."})


@app.route("/api/dashboard/stream")
def api_dashboard_stream():
    tok = request.args.get("token")
    if not tok or not verify_auth_token(tok):
        return jsonify({"error": "Authentication required"}), 401

    q = Queue()
    with sse_lock:
        sse_clients.append(q)

    def generate():
        try:
            yield "event: connected\ndata: {}\n\n"
            while True:
                try:
                    message = q.get(timeout=30)
                    yield message
                except Exception:
                    yield ": heartbeat\n\n"
        except GeneratorExit:
            pass
        finally:
            with sse_lock:
                if q in sse_clients:
                    sse_clients.remove(q)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"}
    )


# ---------------------------------------------------------------------------
# PDF Export (token required)
# ---------------------------------------------------------------------------
@app.route("/api/export/pdf")
@token_required
def api_export_pdf():
    from fpdf import FPDF

    date_from = request.args.get("from", date.today().isoformat())
    date_to = request.args.get("to", date_from)

    db = get_db()
    rows = db.execute("""
        SELECT c.*, v.name as visitor_name, v.company, v.phone
        FROM checkins c JOIN visitors v ON v.id = c.visitor_id
        WHERE c.date_stamp >= ? AND c.date_stamp <= ?
        ORDER BY c.date_stamp ASC, c.time_in ASC
    """, (date_from, date_to)).fetchall()

    pdf = FPDF(orientation="L", unit="mm", format="Letter")
    pdf.set_auto_page_break(auto=True, margin=15)

    cols = [
        ("#", 7), ("Name", 30), ("Company", 28), ("Phone", 22), ("Site", 28),
        ("Est.", 12), ("Reason", 22), ("Time In", 14), ("UAR #", 14),
        ("Time Out", 14), ("Tkt Clsd", 13), ("Alarms", 13),
        ("Alarm Detail", 20), ("Notes", 24), ("Done", 10),
    ]

    dates_in_range = set(r["date_stamp"] for r in rows)
    if not dates_in_range:
        dates_in_range = {date_from}

    for d in sorted(dates_in_range):
        day_rows = [r for r in rows if r["date_stamp"] == d]
        pdf.add_page()

        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Visitor List", ln=True, align="C")

        try:
            display_date = datetime.strptime(d, "%Y-%m-%d").strftime("%B %d, %Y")
        except Exception:
            display_date = d
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, display_date, ln=True, align="C")
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 6.5)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(241, 245, 249)
        for name, w in cols:
            pdf.cell(w, 7, name, border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(0, 0, 0)
        for i, r in enumerate(day_rows, 1):
            pdf.set_fill_color(240, 240, 245) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            vals = [
                str(i), (r["visitor_name"] or "")[:20], (r["company"] or "")[:18],
                (r["phone"] or "")[:15], (r["site"] or "")[:18],
                (r["estimated_time"] or "")[:7], (r["reason"] or "")[:15],
                r["time_in"] or "", (r["uar_number"] or "")[:10],
                r["time_out"] or "", (r["ticket_closed"] or ""),
                (r["alarms_clear"] or ""), (r["alarm_details"] or "")[:15],
                (r["notes"] or "")[:18], "Y" if r["complete"] else "",
            ]
            for (_, w), val in zip(cols, vals):
                pdf.cell(w, 6, val, border=1, fill=True)
            pdf.ln()

        if not day_rows:
            pdf.set_font("Helvetica", "I", 10)
            pdf.cell(0, 12, "No visits recorded for this date.", ln=True, align="C")

        pdf.ln(4)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, f"Total visits: {len(day_rows)}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)

    filename = f"visitor_log_{date_from}.pdf" if date_from == date_to else f"visitor_log_{date_from}_to_{date_to}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------------------------------------------------------------------------
# Routes — Admin (token required)
# ---------------------------------------------------------------------------
@app.route("/api/admin/visitors")
@token_required
def api_admin_visitors():
    db = get_db()
    rows = db.execute("SELECT * FROM visitors ORDER BY approved ASC, name ASC").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/visitors", methods=["POST"])
@token_required
def api_admin_add_visitor():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    company = (data.get("company") or "").strip()
    phone = (data.get("phone") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not company:
        return jsonify({"error": "Company is required"}), 400

    db = get_db()
    try:
        db.execute("INSERT INTO visitors (name, company, phone, approved) VALUES (?, ?, ?, 1)", (name, company, phone))
        db.execute("INSERT OR IGNORE INTO companies (name) VALUES (?)", (company,))
        db.commit()
        audit("add_visitor", f"{name} ({company})", "staff")
        return jsonify({"success": True})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Visitor already exists with that name and company"}), 409


@app.route("/api/admin/visitors/<int:vid>", methods=["DELETE"])
@token_required
def api_admin_delete_visitor(vid):
    db = get_db()
    visitor = db.execute("SELECT name, company FROM visitors WHERE id = ?", (vid,)).fetchone()
    db.execute("DELETE FROM visitors WHERE id = ?", (vid,))
    db.commit()
    if visitor:
        audit("delete_visitor", f"{visitor['name']} ({visitor['company']})", "staff")
    return jsonify({"success": True})


@app.route("/api/admin/companies")
@token_required
def api_admin_companies():
    db = get_db()
    rows = db.execute("""
        SELECT c.name,
               (SELECT COUNT(*) FROM visitors v WHERE v.company = c.name) as visitor_count,
               (SELECT COUNT(*) FROM company_sites cs WHERE cs.company = c.name) as site_count
        FROM companies c ORDER BY c.name
    """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/companies", methods=["POST"])
@token_required
def api_admin_add_company():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Company name is required"}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO companies (name) VALUES (?)", (name,))
        db.commit()
        audit("add_company", name, "staff")
        return jsonify({"success": True})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Company already exists"}), 409


@app.route("/api/admin/companies/<path:name>", methods=["DELETE"])
@token_required
def api_admin_delete_company(name):
    db = get_db()
    db.execute("DELETE FROM companies WHERE name = ?", (name,))
    db.commit()
    audit("delete_company", name, "staff")
    return jsonify({"success": True})


@app.route("/api/admin/reimport", methods=["POST"])
@token_required
def api_admin_reimport():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("DELETE FROM company_sites; DELETE FROM sites; DELETE FROM reasons; DELETE FROM visitors; DELETE FROM companies;")
    conn.commit()
    conn.close()
    import_from_excel()
    audit("reimport", "Full re-import from Excel", "staff")
    return jsonify({"success": True, "message": "Data re-imported from Excel."})


@app.route("/api/admin/change-password", methods=["POST"])
@token_required
def api_admin_change_password():
    data = request.get_json()
    new_pw = (data.get("password") or "").strip()
    if len(new_pw) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400
    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE config SET value = ? WHERE key = 'staff_password'", (hash_password(new_pw),))
    conn.commit()
    conn.close()
    audit("change_password", "Staff password changed", "staff")
    return jsonify({"success": True, "message": "Password updated."})


@app.route("/api/admin/history")
@token_required
def api_admin_history():
    db = get_db()
    date_filter = request.args.get("date", date.today().isoformat())
    rows = db.execute("""
        SELECT c.*, v.name as visitor_name, v.company, v.phone
        FROM checkins c JOIN visitors v ON v.id = c.visitor_id
        WHERE c.date_stamp = ? ORDER BY c.time_in DESC
    """, (date_filter,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/audit")
@token_required
def api_admin_audit():
    db = get_db()
    limit = min(int(request.args.get("limit", 100)), 500)
    rows = db.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    return jsonify([dict(r) for r in rows])


# ---------------------------------------------------------------------------
# QR Code
# ---------------------------------------------------------------------------
@app.route("/api/qr.png")
def api_qr_png():
    checkin_url = FRONTEND_URL.rstrip("/") + "/"
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(checkin_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="image/png")
    except ImportError:
        return "QR code library not installed", 500


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------
@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    import_from_excel()
    port = int(os.environ.get("PORT", 5000))
    print(f"\n  NCEMC API running on port {port}")
    print(f"  Frontend URL: {FRONTEND_URL}\n")
    app.run(host="0.0.0.0", port=port, debug=True, threaded=True)
